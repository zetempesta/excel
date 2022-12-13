import openpyxl  as xlsx

class Excel:
    Workbook= None
    Sheet = None
    Arquivo = None


    def __init__(self, arquivo, sheet):
        self.Arquivo = arquivo
        self.Workbook = xlsx.load_workbook(filename=arquivo)
        self.Sheet = self.Workbook[sheet]

    def __del__(self):
        self.close_document()


    def open_file(self, arquivo, sheet):
        self.Arquivo = arquivo
        self.Workbook = xlsx.load_workbook(filename=arquivo)
        self.Sheet = self.Workbook[sheet]
    
    def read_cell(self, linha, coluna):
        return self.Sheet.cell(linha,coluna).value
    
    def write_cell(self, linha, coluna, valor):
        self.Sheet.cell(linha,coluna,valor)
        self.Workbook.save(self.Arquivo)
    
    def read_cell_without_save(self, linha, coluna, valor):
        self.Sheet.cell(linha,coluna,valor)

    def save(self):
        self.Workbook.save(self.Arquivo)
        
    def close_document(self):
        self.Workbook.close

    def read_cell_commit(self, arquivo, sheet, linha, coluna):
        self.open_file(arquivo=arquivo,sheet=sheet)
        valor = self.read_cell(linha=linha, coluna=coluna)
        return valor

    def write_cell_commit(self, arquivo, sheet, linha, coluna, valor):
        self.open_file(arquivo=arquivo,sheet=sheet)
        self.write_cell(linha,coluna,valor)

